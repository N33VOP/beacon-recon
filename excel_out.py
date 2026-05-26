"""
One place that writes Lisa's workbook, cleanly formatted:
bold frozen header, sensible column widths, and rows colour-coded by severity.
Used by both the app and the folder script so the output is identical.
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Fills are dark text on a soft tint -> readable in Excel regardless of theme.
SEV_FILL = {
    "DROPPED LINE":      "F8C9C9",
    "QTY MISMATCH":      "FBD9B5",
    "UNKNOWN PO":        "FBD9B5",
    "PRICE DRIFT":       "FCEC9E",
    "CURRENCY — REVIEW": "DCD2F5",
    "DATE SLIP":         "C9DEF8",
    "NO PRICE SHOWN":    "E6E6E6",
    "NEEDS REVIEW":      "E6E6E6",
    "OK":                "CDE8C5",
}
HEADER_FILL = "2F3B52"
THIN = Side(style="thin", color="D0D0D0")


def _format(ws, color_by_issue=True):
    headers = [c.value for c in ws[1]]
    # header styling
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # column widths from content
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        longest = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(max(longest + 2, 12), 70)

    # row colour by severity + light borders + wrap the detail column
    icol = headers.index("issue") + 1 if "issue" in headers else None
    dcol = headers.index("detail") + 1 if "detail" in headers else None
    for row in ws.iter_rows(min_row=2):
        fill = SEV_FILL.get(row[icol - 1].value) if (color_by_issue and icol) else None
        for cell in row:
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=(dcol and cell.column == dcol))
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)


def write_workbook(path_or_buffer, action, results, scans=None, errors=None):
    scans = scans or []
    errors = errors or []
    with pd.ExcelWriter(path_or_buffer, engine="openpyxl") as xl:
        action.to_excel(xl, sheet_name="Action Needed", index=False)
        results.to_excel(xl, sheet_name="All Results", index=False)
        pd.DataFrame({"scanned_file": scans}).to_excel(xl, sheet_name="Unreadable Files", index=False)
        pd.DataFrame(errors if errors else [{"file": "", "error": ""}]).to_excel(
            xl, sheet_name="Errors", index=False)
        _format(xl.book["Action Needed"])
        _format(xl.book["All Results"])
        _format(xl.book["Unreadable Files"], color_by_issue=False)
        _format(xl.book["Errors"], color_by_issue=False)